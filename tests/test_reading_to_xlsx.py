import io
import zipfile
from openpyxl import load_workbook
from tests.test_archiving_to_zip import ZIP_PATH

def test_reading_to_xlsx():
    # словарь с ожидаемыми значениями: ключ = (строка, колонка), значение = что ожидаем
    expected_values = {(2, 1): "Austen Russel"}

    with zipfile.ZipFile(ZIP_PATH) as zip_file:  # открываем архив и присваиваем ему имя переменной'zip_file'
        with zip_file.open("1mb.xlsx") as xlsx_file:  # открываем файл в архиве и присваиваем ему имя переменной 'xlsx_file'
            workbook = load_workbook(io.BytesIO(xlsx_file.read()))
            sheet = workbook.active

            for (row_index, col_index), expected_value in expected_values.items():
                actual_value = sheet.cell(row=row_index, column=col_index).value
                print(f"Ячейка ({row_index},{col_index}) содержит: '{actual_value}'")  # выводим значение в кавычках

                # проверка через if / else
                if actual_value == expected_value:
                    print(f"✅ Значение соответствует ожидаемому: '{expected_value}'")
                else:
                    print(f"❌ Значение не соответствует! Ожидалось: '{expected_value}', найдено: '{actual_value}'")
                    raise Exception(f"Ячейка ({row_index},{col_index}) не соответствует ожидаемому значению!")
